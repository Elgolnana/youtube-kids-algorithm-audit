"""
전체실험_테스트.py — 전체 실험용 (시청 시간 5분)
로그인 1회 → P1수집 → 오염 → P2수집 → AI개입 → P3수집 → 분석

사용법:
  GROQ_API_KEY="gsk_..." python3 -u "실행 코드/전체실험.py"

재시작:
  중간에 튕겨도 완료된 단계는 자동으로 건너뜀
"""

import asyncio, json, os, re, sys
from pathlib import Path
from urllib.parse import urlparse, parse_qs

# ── 경로 ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
RESULT_DIR   = SCRIPT_DIR.parent / "실행결과"
PROGRESS_FILE = RESULT_DIR / "실험_진행상황.json"

PROFILES     = ["방관형", "보통형", "적극형"]
PERSONAS     = {"방관형": "방관형", "보통형": "보통형", "적극형": "적극형"}
SEED_VIDEO_ID = "p2bfp5dt1Yg"   # 비디오페이지 수집 기준 영상 (중립)
WATCH_SECONDS = 10               # 오염용 영상 시청 시간(초) — 테스트용 10초
SIGNAL_LOGIN  = "/tmp/ytk_login.signal"
SIGNAL_START  = "/tmp/ytk_start.signal"

PHASE_FOLDER = {
    "p1": RESULT_DIR / "1단계_P1수집",
    "p2": RESULT_DIR / "3단계_P2수집",
    "p3": RESULT_DIR / "5단계_P3수집",
}

# ── 진행 상황 추적 ────────────────────────────────────────────────────────────

def load_progress() -> dict:
    if PROGRESS_FILE.exists():
        return json.loads(PROGRESS_FILE.read_text(encoding="utf-8"))
    return {"completed": []}

def mark_done(stage: str):
    p = load_progress()
    if stage not in p["completed"]:
        p["completed"].append(stage)
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    PROGRESS_FILE.write_text(json.dumps(p, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  ✅ [{stage}] 완료 기록")

def is_done(stage: str) -> bool:
    return stage in load_progress()["completed"]


# ── 신호 대기 ─────────────────────────────────────────────────────────────────

async def wait_signal(path: str, label: str):
    if os.path.exists(path):
        os.remove(path)
    print(f"\n[대기] Claude에게 '{label}' 라고 말해주세요.")
    while not os.path.exists(path):
        await asyncio.sleep(1)
    os.remove(path)
    print(f"  ✓ 신호 수신 — {label}")


# ── 공통 헬퍼 ────────────────────────────────────────────────────────────────

async def click_profile(page, name: str) -> bool:
    # 실제 화면 프로필명: 키즈→"키즈 1", 키즈2→"키즈 2", 키즈3→"키즈 3" 가능
    aliases = {"방관형": ["방관형"], "보통형": ["보통형"], "적극형": ["적극형"]}
    names_to_try = aliases.get(name, [name])

    for n in names_to_try:
        try:
            loc = page.locator(f'css=* >> text="{n}"')
            if await loc.count() > 0 and await loc.first.is_visible(timeout=1000):
                await loc.first.click()
                await asyncio.sleep(3)
                return True
        except Exception:
            pass

    try:
        loc = page.locator(f'css=* >> text="{name}"')
        if await loc.count() > 0 and await loc.first.is_visible(timeout=2000):
            await loc.first.click()
            await asyncio.sleep(3)
            return True
    except Exception:
        pass

    for loc in [
        page.get_by_text(name, exact=True),
        page.locator(f"[aria-label='{name}']"),
        page.locator(f"[title='{name}']"),
        page.get_by_role("button", name=name),
    ]:
        try:
            if await loc.first.is_visible(timeout=1500):
                await loc.first.click()
                await asyncio.sleep(3)
                return True
        except Exception:
            pass

    rect = await page.evaluate("""
        (name) => {
            function ancestor(el) {
                let cur = el;
                while (cur) {
                    const tag = cur.tagName || '';
                    const role = (cur.getAttribute && cur.getAttribute('role')) || '';
                    const ti = cur.getAttribute && cur.getAttribute('tabindex');
                    if (tag==='BUTTON'||role==='button'||tag==='A'||(ti!==null&&ti!=='-1')) return cur;
                    cur = cur.parentElement;
                }
                return el;
            }
            function search(root) {
                for (const el of root.querySelectorAll('*')) {
                    const vals = [el.innerText, el.textContent,
                        el.getAttribute&&el.getAttribute('aria-label'),
                        el.getAttribute&&el.getAttribute('title'),
                        el.getAttribute&&el.getAttribute('alt'),
                    ].filter(Boolean).map(v => v.trim());
                    if (vals.some(v => v === name)) {
                        const t = ancestor(el);
                        const r = t.getBoundingClientRect();
                        if (r.width > 0 && r.height > 0) return {x: r.x+r.width/2, y: r.y+r.height/2};
                    }
                    if (el.shadowRoot) { const f = search(el.shadowRoot); if (f) return f; }
                }
                return null;
            }
            return search(document);
        }
    """, name)
    if rect:
        await page.mouse.click(rect["x"], rect["y"])
        await asyncio.sleep(3)
        return True
    return False

async def dismiss_popups(page):
    for sel in ["button[aria-label*='확인']", "ytk-alert-dialog-renderer button"]:
        try:
            btn = page.locator(sel).first
            if await btn.is_visible(timeout=500):
                await btn.click()
        except Exception:
            pass

async def go_home(page, profile_name: str):
    """홈 하트 버튼 클릭으로 복귀 (SPA 네비게이션)"""
    try:
        # 1. 프로필 화면 감지 (비디오 요소 없음 = 프로필 선택 화면)
        try:
            has_video = await page.evaluate("() => !!document.querySelector('video')")
            if not has_video:
                # 프로필 화면 감지 — 자동으로 프로필 클릭
                await click_profile(page, profile_name)
                await asyncio.sleep(3)
        except Exception:
            pass

        # 2. 네비게이션 바에서 가장 왼쪽의 버튼(하트) 찾아 클릭
        clicked = await page.evaluate("""
            () => {
                function findHomeButton(root) {
                    let leftmost = null;
                    let minX = Infinity;

                    for (const el of root.querySelectorAll('button,[role="button"]')) {
                        const r = el.getBoundingClientRect();
                        // 가시 영역이고 위쪽에 있는 버튼 중 가장 왼쪽
                        if (r.width > 0 && r.height > 0 && r.top < 100 && r.left < minX) {
                            minX = r.left;
                            leftmost = el;
                        }
                    }

                    if (leftmost) {
                        leftmost.click();
                        return true;
                    }

                    for (const el of root.querySelectorAll('*')) {
                        if (el.shadowRoot && findHomeButton(el.shadowRoot)) return true;
                    }
                    return false;
                }
                return findHomeButton(document);
            }
        """)

        if not clicked:
            # 폴백: 뒤로가기 버튼 클릭
            try:
                await page.keyboard.press("Escape")
            except Exception:
                pass

        await asyncio.sleep(2)
    except Exception as e:
        print(f"  [{profile_name}] go_home 오류: {e}")


# ── JavaScript ────────────────────────────────────────────────────────────────

EXTRACT_JS = """
(maxItems) => {
    const seen=new Set(),results=[];
    function vid(href){const m=(href||'').match(/[?&]v=([^&]+)/);return m?m[1]:null;}
    function search(root){
        for(const link of root.querySelectorAll('a[href*="/watch"]')){
            if(results.length>=maxItems)break;
            const videoId=vid(link.href||link.getAttribute('href'));
            if(!videoId||seen.has(videoId))continue;
            let title=link.getAttribute('aria-label')||link.getAttribute('title')||'';
            if(!title){for(const s of link.querySelectorAll('span,p,h3,div')){
                const t=(s.innerText||s.textContent||'').trim();if(t&&t.length>2){title=t;break;}
            }}
            const rect=link.getBoundingClientRect();
            if(rect.width===0&&rect.height===0)continue;
            seen.add(videoId);
            const img=link.querySelector('img');
            results.push({video_id:videoId,title:title.trim()||'(제목 없음)',
                url:'https://www.youtubekids.com/watch?v='+videoId,
                thumbnail:(img&&(img.src||img.getAttribute('src')))||''});
        }
        for(const el of root.querySelectorAll('*'))if(el.shadowRoot)search(el.shadowRoot);
    }
    search(document);return results;
}
"""

FIND_MENU_JS = """
(videoId) => {
    function search(root) {
        for (const link of root.querySelectorAll('a[href*="' + videoId + '"]')) {
            const lr=link.getBoundingClientRect();if(lr.width===0)continue;
            let el=link;
            for(let i=0;i<8;i++){
                el=el.parentElement;if(!el)break;
                for(const btn of el.querySelectorAll('button,[role="button"]')){
                    const label=(btn.getAttribute('aria-label')||'').toLowerCase();
                    const br=btn.getBoundingClientRect();
                    if(br.width===0||br.width>60)continue;
                    if(label.includes('more')||label.includes('더보기')||
                       label.includes('옵션')||label.includes('menu'))
                        return{x:br.x+br.width/2,y:br.y+br.height/2};
                }
            }
            for(const btn of document.querySelectorAll('button,[role="button"]')){
                const br=btn.getBoundingClientRect();
                if(br.width===0||br.width>50||br.height>50)continue;
                if(Math.abs(br.top-lr.top)<100&&br.left>lr.right-100)
                    return{x:br.x+br.width/2,y:br.y+br.height/2};
            }
        }
        for(const el of root.querySelectorAll('*'))
            if(el.shadowRoot){const r=search(el.shadowRoot);if(r)return r;}
        return null;
    }
    return search(document);
}
"""


# ══════════════════════════════════════════════════════════════════════════════
# 단계 1·3·5: 피드 수집 (P1 / P2 / P3)
# ══════════════════════════════════════════════════════════════════════════════

def _phase_folder(phase: str) -> str:
    return {"p1":"1단계_P1수집","p2":"3단계_P2수집","p3":"5단계_P3수집"}[phase]

def feed_path(phase: str, profile: str) -> Path:
    base = RESULT_DIR / _phase_folder(phase)
    base.mkdir(parents=True, exist_ok=True)
    return base / f"{phase}_feed_{profile}.json"

def screenshot_path(phase: str, profile: str, scene: str) -> Path:
    base = RESULT_DIR / _phase_folder(phase) / "스크린샷"
    base.mkdir(parents=True, exist_ok=True)
    return base / f"{phase}_{scene}_{profile}.png"

async def collect_one_profile(page, profile_name: str, phase: str):
    """3단계_오염 후 피드수집.py 의 collect_profile_feed 와 동일한 로직"""
    print(f"\n  [{profile_name}] 피드 수집 시작")

    # 홈으로 먼저 이동 (차단 후 재생불가 화면 등 이상 상태 대비)
    await go_home(page, profile_name)
    await asyncio.sleep(15)  # 차단 후 피드 재로드 대기
    await dismiss_popups(page)

    # 홈피드 — 피드 링크 없으면 프로필 재클릭
    print(f"  [{profile_name}] 홈페이지...")
    has_feed = await page.evaluate("""
        () => {
            function check(root) {
                for (const a of root.querySelectorAll('a[href*="/watch"]')) {
                    if (a.getBoundingClientRect().width > 0) return true;
                }
                for (const el of root.querySelectorAll('*'))
                    if (el.shadowRoot && check(el.shadowRoot)) return true;
                return false;
            }
            return check(document);
        }
    """)
    if not has_feed:
        print(f"  [{profile_name}] 피드 없음 → 프로필 재클릭")
        await click_profile(page, profile_name)
        await asyncio.sleep(4)
        await dismiss_popups(page)

    for _ in range(3):
        await page.evaluate("window.scrollBy(0, 600)")
        await asyncio.sleep(1.5)
    home_videos = await page.evaluate(EXTRACT_JS, 10)
    await page.screenshot(path=str(screenshot_path(phase, profile_name, "home")), full_page=True)
    print(f"  [{profile_name}] 홈 추천 {len(home_videos)}개")

    # 비디오페이지 — goto 후 프로필 재선택 필요 시 처리
    print(f"  [{profile_name}] 비디오페이지...")
    await page.goto(f"https://www.youtubekids.com/watch?v={SEED_VIDEO_ID}",
                    wait_until="domcontentloaded", timeout=30000)
    await asyncio.sleep(5)
    await dismiss_popups(page)

    has_video = await page.evaluate("() => !!document.querySelector('video')")
    if not has_video:
        print(f"  [{profile_name}] 프로필 재선택 필요 → 클릭")
        await click_profile(page, profile_name)
        await asyncio.sleep(5)
        await dismiss_popups(page)

    for _ in range(3):
        await page.evaluate("""
            () => {
                function tryScroll(root) {
                    for (const el of root.querySelectorAll('*')) {
                        const s = window.getComputedStyle(el);
                        if ((s.overflowY==='auto'||s.overflowY==='scroll') &&
                            el.scrollHeight > el.clientHeight+10 && el.clientHeight>150) {
                            el.scrollTop += 600; return true;
                        }
                        if (el.shadowRoot && tryScroll(el.shadowRoot)) return true;
                    }
                    window.scrollBy(0, 600);
                }
                tryScroll(document);
            }
        """)
        await asyncio.sleep(1.5)
    raw = await page.evaluate(EXTRACT_JS, 20)
    vp_videos = [v for v in raw if v["video_id"] != SEED_VIDEO_ID][:10]
    # 스크롤 초기화 후 4장 찍어 이어붙이기
    import io as _io
    from PIL import Image as _PILImage
    await page.evaluate("""
        () => {
            function resetScroll(root) {
                for (const el of root.querySelectorAll('*')) {
                    if (el.scrollHeight > el.clientHeight+50 && el.clientHeight>200) {
                        el.scrollTop = 0; return true;
                    }
                    if (el.shadowRoot && resetScroll(el.shadowRoot)) return true;
                }
                return false;
            }
            resetScroll(document);
        }
    """)
    await asyncio.sleep(0.5)
    chunks = []
    for _ in range(4):
        buf = await page.screenshot(full_page=False)
        chunks.append(_PILImage.open(_io.BytesIO(buf)))
        await page.evaluate("""
            () => {
                function tryScroll(root) {
                    for (const el of root.querySelectorAll('*')) {
                        if (el.scrollHeight > el.clientHeight+50 && el.clientHeight>200) {
                            el.scrollTop += 700; return true;
                        }
                        if (el.shadowRoot && tryScroll(el.shadowRoot)) return true;
                    }
                    window.scrollBy(0, 700);
                }
                tryScroll(document);
            }
        """)
        await asyncio.sleep(1.5)
    total_h = sum(img.height for img in chunks)
    combined = _PILImage.new("RGB", (chunks[0].width, total_h))
    y_off = 0
    for img in chunks:
        combined.paste(img, (0, y_off)); y_off += img.height
    combined.save(str(screenshot_path(phase, profile_name, "videopage")))
    print(f"  [{profile_name}] 비디오페이지 추천 {len(vp_videos)}개")

    result = {"phase": phase, "profile": profile_name,
              "home": home_videos, "video_page": vp_videos}
    feed_path(phase, profile_name).write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return result

def generate_screenshot_pdf(phase: str):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Image, Paragraph, Spacer
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from PIL import Image as PILImage

        fp = Path.home()/"Library/Fonts/NanumSquare_acR.ttf"
        fb = Path.home()/"Library/Fonts/NanumSquare_acB.ttf"
        if fp.exists():
            pdfmetrics.registerFont(TTFont("KR", str(fp)))
            pdfmetrics.registerFont(TTFont("KR-B", str(fb)))
            kr, kr_b = "KR", "KR-B"
        else:
            kr, kr_b = "Helvetica", "Helvetica-Bold"

        folder = RESULT_DIR / _phase_folder(phase) / "스크린샷"
        if not folder.exists():
            return
        pdf_path = RESULT_DIR / _phase_folder(phase) / f"{phase}_스크린샷.pdf"
        label = {"p1":"P1 (오염 전)","p2":"P2 (오염 후)","p3":"P3 (개입 후)"}[phase]
        story = [
            Paragraph(f"YouTube Kids 피드 스크린샷 — {label}",
                      ParagraphStyle("T", fontName=kr_b, fontSize=16, spaceAfter=8)),
            Spacer(1, 0.5*cm),
        ]
        for profile in PROFILES:
            story.append(Paragraph(f"프로필: {profile}",
                                   ParagraphStyle("H2", fontName=kr_b, fontSize=13, spaceAfter=4)))
            for scene, slabel in [("home","홈피드"),("videopage","비디오페이지")]:
                img_p = folder / f"{phase}_{scene}_{profile}.png"
                if img_p.exists():
                    story.append(Paragraph(slabel,
                                           ParagraphStyle("H3", fontName=kr, fontSize=11, spaceAfter=4)))
                    pil = PILImage.open(img_p)
                    orig_w, orig_h = pil.size
                    pil.close()
                    max_w = 17 * cm
                    display_h = min(max_w * orig_h / orig_w, 24 * cm)
                    display_w = display_h * orig_w / orig_h
                    story.append(Image(str(img_p), width=display_w, height=display_h))
                    story.append(Spacer(1, 0.4*cm))
        doc = SimpleDocTemplate(str(pdf_path), pagesize=A4,
                                leftMargin=1*cm, rightMargin=1*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        doc.build(story)
        print(f"  PDF 저장: {pdf_path.name}")
    except Exception as e:
        print(f"  [!] PDF 생성 실패: {e}")

async def stage_collect(pages: list, phase: str):
    label = {"p1":"P1 수집","p2":"P2 수집","p3":"P3 수집"}[phase]
    print(f"\n{'─'*60}\n[{label}] 3개 프로필 동시 수집\n{'─'*60}")
    await asyncio.gather(*[
        collect_one_profile(pages[i], PROFILES[i], phase)
        for i in range(len(PROFILES))
    ], return_exceptions=True)
    generate_screenshot_pdf(phase)
    mark_done(phase)


# ══════════════════════════════════════════════════════════════════════════════
# 단계 2: 오염 (시드 영상 시청)
# ══════════════════════════════════════════════════════════════════════════════

def load_seed_videos() -> list:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("openpyxl 미설치: pip install openpyxl")
    candidates = [
        SCRIPT_DIR / "시드영상_목록.xlsx",
        SCRIPT_DIR.parent / "시드영상_목록.xlsx",
        SCRIPT_DIR.parent / "과제 안내 및 연구 설계" / "data" / "시드영상_목록.xlsx",
    ]
    xlsx = next((p for p in candidates if p.exists()), None)
    if not xlsx:
        raise SystemExit("시드영상_목록.xlsx 를 찾을 수 없습니다.")
    wb = openpyxl.load_workbook(xlsx)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    title_col = next((i for i,h in enumerate(headers) if h and "제목" in str(h)), None)
    url_col   = next((i for i,h in enumerate(headers) if h and "url" in str(h).lower()), None)
    if url_col is None:
        raise SystemExit("시드영상_목록.xlsx에 URL 컬럼이 없습니다.")
    type_col = next((i for i,h in enumerate(headers) if h and "유형" in str(h)), None)
    from collections import defaultdict
    by_type = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[url_col]
        if not url: continue
        qs = parse_qs(urlparse(str(url)).query)
        if "v" in qs:
            vid = qs["v"][0]
        else:
            m = re.search(r"youtu\.be/([^?&]+)", str(url))
            vid = m.group(1) if m else None
        if not vid: continue
        title = str(row[title_col]) if title_col is not None and row[title_col] else vid
        type_key = str(row[type_col]) if type_col is not None and row[type_col] else "기타"
        by_type[type_key].append((vid, title))
    # 테스트: 유형별 최대 5개씩 선택 (최종 실험 시 이 부분 제거)
    import random as _random
    videos = []
    for t, vlist in by_type.items():
        sampled = _random.sample(vlist, min(7, len(vlist)))
        videos.extend(sampled)
    print(f"  시드 영상 {len(videos)}개 로드 (유형별 무작위 7개)")
    return videos

def stain_log_path(profile: str) -> Path:
    base = RESULT_DIR / "2단계_오염"
    base.mkdir(parents=True, exist_ok=True)
    return base / f"stain_log_{profile}.json"

def done_video_ids(profile: str) -> set:
    p = stain_log_path(profile)
    if not p.exists(): return set()
    try:
        return {r["video_id"] for r in json.loads(p.read_text(encoding="utf-8")).get("results",[])
                if r.get("status")=="done"}
    except Exception: return set()

async def watch_video(page, profile_name: str, video_id: str, title: str,
                      idx: int, total: int) -> dict:
    """URL로 직접 이동 → 시청 (2단계_오염 시키기.py 방식)"""
    url = f"https://www.youtubekids.com/watch?v={video_id}"
    print(f"  [{profile_name}] [{idx}/{total}] {title[:40]}")

    # URL로 직접 이동 (1회 재시도)
    for attempt in range(2):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=40000)
            break
        except Exception as e:
            if attempt == 0:
                print(f"  [{profile_name}] goto 재시도... ({e})")
                await asyncio.sleep(3)
            else:
                print(f"  [{profile_name}] goto 실패 → 건너뜀")
                return {"video_id": video_id, "title": title, "watched_seconds": 0, "status": "error"}

    await asyncio.sleep(4)
    await dismiss_popups(page)

    # 프로필 화면이면 자동 재선택 후 URL 재이동
    has_video_el = await page.evaluate("() => !!document.querySelector('video')")
    if not has_video_el:
        print(f"  [{profile_name}] 프로필 화면 감지 → 재선택")
        await click_profile(page, profile_name)
        await asyncio.sleep(3)
        await page.goto(url, wait_until="domcontentloaded", timeout=40000)
        await asyncio.sleep(4)
        await dismiss_popups(page)

    # 재생
    for sel in ["video", "button[aria-label*='재생']"]:
        try:
            el = page.locator(sel).first
            if await el.is_visible(timeout=2000): await el.click(); break
        except Exception: pass

    # 시청
    watched = 0
    while watched < WATCH_SECONDS:
        await asyncio.sleep(15)
        watched += 15
        await dismiss_popups(page)
        try:
            paused = await page.evaluate(
                "() => {const v=document.querySelector('video');return v?v.paused:true;}"
            )
            if paused:
                await page.evaluate("() => {const v=document.querySelector('video');if(v)v.play();}")
        except Exception: pass
        filled = int(20*watched/WATCH_SECONDS)
        print(f"  [{profile_name}] [{'█'*filled+'░'*(20-filled)}] {watched}s/{WATCH_SECONDS}s", end="\r", flush=True)

    print(f"\n  [{profile_name}] ✓ 완료: {title[:40]}")

    return {"video_id": video_id, "title": title,
            "watched_seconds": WATCH_SECONDS, "status": "done"}

async def profile_guardian(page, profile_name: str, stop_event: asyncio.Event):
    """2단계_오염 시키기.py 와 동일 — watch 페이지에서 프로필 화면 뜨면 자동 재클릭"""
    while not stop_event.is_set():
        try:
            s = await page.evaluate(
                "() => ({url: location.href, hasVideo: !!document.querySelector('video')})"
            )
            if "youtubekids.com/watch" in s["url"] and not s["hasVideo"]:
                print(f"\n  [{profile_name}] 가디언: 프로필 화면 감지 → 자동 클릭")
                await click_profile(page, profile_name)
        except Exception:
            pass
        await asyncio.sleep(1)

async def contaminate_profile(page, profile_name: str, seed_videos: list):
    done_ids = done_video_ids(profile_name)
    remaining = [(v,t) for v,t in seed_videos if v not in done_ids]
    if not remaining:
        print(f"  [{profile_name}] 이미 완료됨 — 건너뜀")
        return
    print(f"  [{profile_name}] {len(remaining)}개 시청 시작")
    log = stain_log_path(profile_name)
    existing = json.loads(log.read_text(encoding="utf-8")).get("results",[]) if log.exists() else []
    stop_event = asyncio.Event()
    guardian = asyncio.create_task(profile_guardian(page, profile_name, stop_event))
    new_results = []
    for i, (vid, title) in enumerate(seed_videos, 1):
        if vid in done_ids: continue
        try:
            result = await watch_video(page, profile_name, vid, title, i, len(seed_videos))
        except Exception as e:
            print(f"  [{profile_name}] 오류: {e}")
            result = {"video_id":vid,"title":title,"watched_seconds":0,"status":"error"}
        new_results.append(result)
        log.write_text(json.dumps({"profile":profile_name,"total":len(seed_videos),
                                   "results":existing+new_results},
                                  ensure_ascii=False, indent=2), encoding="utf-8")
    stop_event.set()
    await guardian
    print(f"  [{profile_name}] 오염 완료")

async def stage_contaminate(pages: list):
    print(f"\n{'─'*60}\n[2단계: 오염] 시드 영상 시청\n{'─'*60}")
    seed_videos = load_seed_videos()
    await asyncio.gather(*[
        contaminate_profile(pages[i], PROFILES[i], seed_videos)
        for i in range(len(PROFILES))
    ], return_exceptions=True)
    mark_done("contaminate")


# ══════════════════════════════════════════════════════════════════════════════
# 단계 4: AI 개입 (Judge + Executor)
# ══════════════════════════════════════════════════════════════════════════════

LOG_JUDGE = RESULT_DIR / "4단계_AI개입"
LOG_JUDGE.mkdir(parents=True, exist_ok=True)


# ── 기초 차단 키워드 (페르소나별 고정) ─────────────────────────────────────────
BASE_KEYWORDS = {
    "방관형":  [],
    "보통형": ["19금", "성인용", "이혼", "부부싸움", "피어싱", "시청주의", "경고",
              "런닝맨", "아는형님", "이혼숙려캠프", "최강야구",
              "체인소맨", "귀멸의칼날", "JTBC"],
    "적극형": ["19금", "성인용", "이혼", "부부싸움", "피어싱", "시청주의", "경고",
              "런닝맨", "아는형님", "이혼숙려캠프", "최강야구",
              "체인소맨", "귀멸의칼날", "주술회전", "진격의거인", "JTBC",
              "사춘기", "화장", "다이어트", "성형", "공포", "귀신",
              "술", "담배", "연애", "남친", "여친", "엄마 몰래", "아빠한테"],
}

async def run_judge_agent(profile_name: str, persona: str, p2_feed: dict) -> dict:
    if persona == "방관형":
        criteria = {"keywords": [], "block_type": "none", "rationale": "방관형은 차단하지 않음"}
        (LOG_JUDGE / f"criteria_{profile_name}.json").write_text(
            json.dumps(criteria, ensure_ascii=False, indent=2), encoding="utf-8")
        return criteria

    from groq import Groq
    client = Groq(api_key=os.environ.get("GROQ_API_KEY", ""))
    all_videos = list({v["video_id"]: v
                       for v in p2_feed["home"] + p2_feed["video_page"]}.values())
    def clean_title(t: str) -> str:
        import re as _re
        t = _re.sub(r'^동영상\s+', '', t.strip())
        t = _re.sub(r'\s*\d+분\s*\d*초?\s*$', '', t).strip()
        return t
    video_list = "\n".join(f"- [{v['video_id']}] {clean_title(v['title'])}" for v in all_videos)
    base_kw = BASE_KEYWORDS.get(profile_name, [])

    persona_guide = {
        "보통형": (
            "보통형 부모 관점에서 다음 두 가지를 추출하세요:\n"
            "1) 성인 예능 프로그램명·출연자명 (런닝맨, 아는형님, 김종국 등 고유명사)\n"
            "2) 청소년 폭력 애니메이션 제목 (체인소맨, 귀멸의칼날 등 고유명사)\n"
            "3) 기초 키워드에 없는 추가 부적절 단어 (19금, 이혼, 피어싱 등 유형의 새로운 단어)\n"
            "block_type은 \"video\"."
        ),
        "적극형": (
            "적극형 부모 관점에서 보통형보다 넓게 다음을 추출하세요:\n"
            "1) 성인 예능 프로그램명·출연자명 고유명사\n"
            "2) 청소년 폭력 애니메이션 제목 고유명사 (주술회전, 진격의거인 포함)\n"
            "3) 기초 키워드에 없는 추가 부적절 단어 (사춘기·화장·공포·연애 유형의 새로운 단어)\n"
            "block_type은 \"channel\"."
        ),
    }[persona]

    print(f"  [Judge Agent] {profile_name} — P2 데이터 분석 중...")

    response = await asyncio.to_thread(
        client.chat.completions.create,
        model="llama-3.3-70b-versatile",
        max_tokens=800,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": "당신은 아동 콘텐츠 안전 전문가입니다. JSON만 출력하세요."},
            {"role": "user", "content": f"""YouTube Kids 오염 후 추천된 영상 목록입니다.
만 5~8세 아동에게 부적합한 영상을 추가로 차단할 키워드를 추출하세요.

페르소나: {persona}
{persona_guide}

[이미 적용된 기초 키워드 — 중복 추출 불필요]
{base_kw}

[오염 후 추천 영상 — {len(all_videos)}개]
{video_list}

출력 형식(JSON만):
{{
  "keywords": ["키워드1", "키워드2", ...],
  "block_type": "video" 또는 "channel",
  "rationale": "한 줄 요약"
}}

반드시 지킬 규칙:
- 고유명사(프로그램명, 출연자명, 애니 제목) 또는 명백히 부적절한 단어만 추출
- 매직·형·가족·학원·병원·경찰·사고·먹방 등 일반 한국어 단어 절대 금지
- 디즈니·핑크퐁·뽀로로·타요·카봇 등 공인 아동 콘텐츠 브랜드 절대 금지
- 3글자 미만 단어 절대 금지"""}
        ]
    )

    criteria = json.loads(response.choices[0].message.content)
    # Judge 키워드 + 기초 키워드 합산
    judge_kw = criteria.get("keywords", [])
    all_kw = list(dict.fromkeys(base_kw + judge_kw))  # 중복 제거, 순서 유지
    criteria["keywords"] = all_kw
    print(f"  [Judge Agent] 기초 {len(base_kw)}개 + 추가 {len(judge_kw)}개 = 총 {len(all_kw)}개")

    (LOG_JUDGE / f"criteria_{profile_name}.json").write_text(
        json.dumps(criteria, ensure_ascii=False, indent=2), encoding="utf-8")
    return criteria

async def _click_block_menu(page, video_id: str, block_type: str, title: str) -> bool:
    await asyncio.sleep(1.5)
    menu_clicked = False
    for text in ["이 동영상 차단","차단","Block"]:
        try:
            loc = page.get_by_text(text, exact=True).first
            if await loc.is_visible(timeout=2000):
                await loc.click()
                menu_clicked = True
                break
        except Exception: pass
    if not menu_clicked: return False
    await asyncio.sleep(0.8)
    if block_type == "channel":
        try:
            loc = page.get_by_text("채널 전체를 차단").first
            if await loc.is_visible(timeout=1500): await loc.click()
        except Exception: pass
        await asyncio.sleep(0.5)
        try:
            btn = page.get_by_role("button", name="차단")
            if await btn.first.is_visible(timeout=1500): await btn.first.click()
        except Exception: pass
        await asyncio.sleep(0.5)
    label = "동영상" if block_type=="video" else "채널"
    print(f"    ✓ [{label} 차단] {title[:45]}")
    return True

async def _channel_block_via_watch(page, profile_name: str, video_id: str, title: str) -> bool:
    video_url = f"https://www.youtubekids.com/watch?v={video_id}"
    used_goto = False

    # 1단계: 피드 링크 클릭 (SPA 이동 — 프로필 유지, go_back()으로 피드 상태 복원)
    clicked = await page.evaluate(f"""
        () => {{
            function s(root){{
                for(const a of root.querySelectorAll('a[href*="{video_id}"]')){{
                    a.scrollIntoView({{behavior:'instant', block:'center'}});
                    if(a.getBoundingClientRect().width > 0){{ a.click(); return true; }}
                }}
                for(const el of root.querySelectorAll('*'))
                    if(el.shadowRoot && s(el.shadowRoot)) return true;
                return false;
            }}
            return s(document);
        }}
    """)

    if clicked:
        await asyncio.sleep(4)
        await dismiss_popups(page)
        has_video = await page.evaluate("() => !!document.querySelector('video')")
        if not has_video:
            print(f"  [{profile_name}] 링크 클릭 후 프로필 리셋 → goto 폴백")
            await page.go_back()
            await asyncio.sleep(2)
            clicked = False

    if not clicked:
        # 2단계: goto 폴백 (프로필 재시도 최대 3회)
        used_goto = True
        print(f"  [{profile_name}] goto 폴백: {title[:40]}")
        video_loaded = False
        for attempt in range(3):
            try:
                await page.goto(video_url, wait_until="domcontentloaded", timeout=30000)
                await asyncio.sleep(6)
                await dismiss_popups(page)
                has_video = await page.evaluate("() => !!document.querySelector('video')")
                if not has_video:
                    print(f"  [{profile_name}] 프로필 화면 → 재선택 ({attempt+1}/3)")
                    await click_profile(page, profile_name)
                    await asyncio.sleep(4)
                    has_video = await page.evaluate("() => !!document.querySelector('video')")
                if has_video:
                    video_loaded = True
                    break
            except Exception as e:
                print(f"  [{profile_name}] goto 오류 ({attempt+1}/3): {e}")
                await asyncio.sleep(2)
        if not video_loaded:
            print(f"  [{profile_name}] 영상 페이지 진입 실패 → 건너뜀: {title[:40]}")
            await go_home(page, profile_name)
            return False

    # 공통: 더보기 메뉴 → 채널 차단
    try:
        box = await page.locator("video").first.bounding_box()
        if box: await page.mouse.move(box["x"]+box["width"]/2, box["y"]+box["height"]/2)
        await asyncio.sleep(0.8)
    except Exception: pass

    coord = await page.evaluate("""
        () => {
            function s(root){
                for(const el of root.querySelectorAll('button,[role="button"]')){
                    const label=(el.getAttribute('aria-label')||'').toLowerCase();
                    const r=el.getBoundingClientRect();
                    if(r.width===0||r.width>80)continue;
                    if(label.includes('more')||label.includes('더보기')||label.includes('메뉴')||label.includes('option'))
                        return{x:r.x+r.width/2,y:r.y+r.height/2};
                    if(el.shadowRoot){const f=s(el.shadowRoot);if(f)return f;}
                }
                return null;
            }
            return s(document);
        }
    """)
    if not coord:
        print(f"  [{profile_name}] 메뉴 버튼 못 찾음 → 건너뜀: {title[:40]}")
        if used_goto:
            await go_home(page, profile_name)
        else:
            await page.go_back()
        await asyncio.sleep(2)
        return False

    await page.mouse.click(coord["x"], coord["y"])
    await asyncio.sleep(1)

    menu_clicked = False
    for text in ["동영상 차단", "이 동영상 차단"]:
        try:
            loc = page.get_by_text(text, exact=True).first
            if await loc.is_visible(timeout=2000):
                await loc.click(); menu_clicked = True; break
        except Exception: pass

    if not menu_clicked:
        print(f"  [{profile_name}] 차단 메뉴 못 찾음 → 건너뜀: {title[:40]}")
        if used_goto:
            await go_home(page, profile_name)
        else:
            await page.go_back()
        await asyncio.sleep(2)
        return False

    await asyncio.sleep(1)
    try:
        loc = page.get_by_text("채널 전체를 차단").first
        if await loc.is_visible(timeout=1500): await loc.click()
    except Exception: pass
    await asyncio.sleep(0.5)

    success = False
    try:
        btn = page.get_by_role("button", name="차단")
        if await btn.first.is_visible(timeout=1500):
            await btn.first.click(); success = True
    except Exception: pass

    if success: print(f"    ✓ [채널 차단] {title[:45]}")

    if used_goto:
        await go_home(page, profile_name)
    else:
        await page.go_back()
    await asyncio.sleep(2)
    return success

async def execute_block(page, profile_name: str, video_id: str, title: str, block_type: str) -> bool:
    import re as _re
    if not _re.match(r'^[A-Za-z0-9_-]{11}$', video_id): return False
    if video_id == SEED_VIDEO_ID:
        print(f"    ⚠ 씨드 영상 채널 차단 예외 처리 — 건너뜀: {title[:45]}")
        return False
    if block_type == "channel":
        return await _channel_block_via_watch(page, profile_name, video_id, title)
    await page.evaluate(f"""
        () => {{
            function sc(root){{
                for(const a of root.querySelectorAll('a[href*="{video_id}"]')){{
                    const r=a.getBoundingClientRect();
                    if(r.width>0){{a.scrollIntoView({{behavior:'instant',block:'center'}});return true;}}
                }}
                for(const el of root.querySelectorAll('*'))if(el.shadowRoot&&sc(el.shadowRoot))return true;
                return false;
            }}
            return sc(document);
        }}
    """)
    await asyncio.sleep(0.5)
    coord = await page.evaluate(FIND_MENU_JS, video_id)
    if not coord: return False
    await page.mouse.click(coord["x"], coord["y"])
    return await _click_block_menu(page, video_id, block_type, title)

async def run_executor(profile_name: str, persona: str, criteria: dict, page) -> dict:
    keywords  = criteria.get("keywords",[])
    block_type = criteria.get("block_type","none")
    action_log = []
    blocked_ids = set()
    print(f"\n  [{profile_name}] Executor 시작 — 키워드 {len(keywords)}개")
    if block_type == "none" or not keywords:
        result = {"profile":profile_name,"persona":persona,"actions":[],"summary":"방관형: 차단 없음"}
        (LOG_JUDGE/f"judgment_{profile_name}.json").write_text(
            json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        return result
    # 홈피드에서 차단 시작 (P2 수집이 비디오페이지에서 끝나므로 홈으로 먼저 이동)
    await go_home(page, profile_name)
    await asyncio.sleep(3)
    await dismiss_popups(page)
    for round_num in range(4):
        await dismiss_popups(page)
        videos = await page.evaluate(EXTRACT_JS, 20)
        print(f"  [{profile_name}] 피드 {len(videos)}개 (라운드 {round_num+1})")
        for video in videos:
            vid = video["video_id"]; title = video["title"]
            if vid in blocked_ids or vid == SEED_VIDEO_ID: continue
            matched = [kw for kw in keywords if len(kw.strip()) >= 3 and kw in title]
            if not matched: continue
            print(f"\n  → [block_{block_type}] {title[:50]}\n     매칭: {matched}")
            success = await execute_block(page, profile_name, vid, title, block_type)
            blocked_ids.add(vid)
            action_log.append({"action":f"block_{block_type}","video_id":vid,
                               "title":title,"matched_keywords":matched,"success":success})
        if round_num < 3:
            for _ in range(3):
                await page.evaluate("window.scrollBy(0,600)")
                await asyncio.sleep(1.5)
    success_count = sum(1 for a in action_log if a.get("success"))
    summary = f"{persona}: {success_count}/{len(action_log)}개 차단"
    print(f"  [{profile_name}] {summary}")
    result = {"profile":profile_name,"persona":persona,"actions":action_log,"summary":summary}
    (LOG_JUDGE/f"judgment_{profile_name}.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result

async def stage_intervene(pages: list):
    print(f"\n{'─'*60}\n[4단계: AI 개입] Judge → Executor\n{'─'*60}")
    # Judge (LLM, 순차)
    criteria_map = {}
    for profile in PROFILES:
        cache = LOG_JUDGE / f"criteria_{profile}.json"
        if cache.exists():
            criteria_map[profile] = json.loads(cache.read_text(encoding="utf-8"))
            print(f"  [{profile}] 캐시된 기준 로드")
        else:
            p2 = json.loads((RESULT_DIR/"3단계_P2수집"/f"p2_feed_{profile}.json")
                            .read_text(encoding="utf-8"))
            criteria_map[profile] = await run_judge_agent(profile, PERSONAS[profile], p2)
    # Executor 전 모든 프로필 홈피드로 이동
    print("\n  홈피드로 이동 중...")
    await asyncio.gather(*[go_home(pages[i], PROFILES[i]) for i in range(len(PROFILES))], return_exceptions=True)
    await asyncio.sleep(3)
    # Executor (병렬)
    print("\n  Executor 차단 시작...")
    await asyncio.gather(*[
        run_executor(PROFILES[i], PERSONAS[PROFILES[i]], criteria_map[PROFILES[i]], pages[i])
        for i in range(len(PROFILES))
    ])
    # 홈으로 복귀
    await asyncio.gather(*[go_home(pages[i], PROFILES[i]) for i in range(len(PROFILES))])
    mark_done("intervene")


# ══════════════════════════════════════════════════════════════════════════════
# 단계 6: 분석
# ══════════════════════════════════════════════════════════════════════════════

def stage_analyze():
    print(f"\n{'─'*60}\n[6단계: 분석]\n{'─'*60}")
    try:
        import pandas as pd
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.font_manager as fm
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.platypus import Image as RLImage
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        fp = Path.home()/"Library/Fonts/NanumSquare_acR.ttf"
        fb = Path.home()/"Library/Fonts/NanumSquare_acB.ttf"
        if fp.exists():
            pdfmetrics.registerFont(TTFont("KR",str(fp)))
            pdfmetrics.registerFont(TTFont("KR-B",str(fb)))
            fm.fontManager.addfont(str(fp))
            plt.rcParams["font.family"] = fm.FontProperties(fname=str(fp)).get_name()
            kr, kr_b = "KR","KR-B"
        else:
            kr, kr_b = "Helvetica","Helvetica-Bold"
        plt.rcParams["axes.unicode_minus"] = False

        out_dir = RESULT_DIR / "6단계_분석"
        out_dir.mkdir(parents=True, exist_ok=True)

        COMMON_DETECTION_KEYWORDS = [
            "체인소맨", "주술회전", "진격의거인", "귀멸의칼날",
            "에렌", "아커만", "리바이", "미카사", "헤이안", "스쿠나",
            "이혼숙려캠프", "이혼", "런닝맨", "아는형님", "JTBC", "스타킹",
            "순풍산부인과", "틱톡", "로블록스",
        ]

        rows = []
        for profile in PROFILES:
            for phase in ["p1","p2","p3"]:
                fp2 = RESULT_DIR / _phase_folder(phase) / f"{phase}_feed_{profile}.json"
                if not fp2.exists(): continue
                data = json.loads(fp2.read_text(encoding="utf-8"))
                home_flagged, home_total = 0, 0
                for scene, slabel in [("home","홈피드"),("video_page","비디오페이지")]:
                    videos = data.get(scene,[])
                    flagged = [v for v in videos if any(k in v.get("title","") for k in COMMON_DETECTION_KEYWORDS)]
                    ratio = len(flagged)/len(videos)*100 if videos else 0
                    rows.append({"profile":profile,"persona":PERSONAS[profile],
                                 "phase":phase,"scene":scene,"scene_label":slabel,
                                 "total":len(videos),"flagged":len(flagged),"ratio":round(ratio,1)})
                    if scene == "home":
                        home_flagged, home_total = len(flagged), len(videos)
                print(f"  [{profile}] {phase.upper()}: 홈 {home_flagged}/{home_total}")

        if not rows:
            print("  [!] 분석할 데이터 없음"); return

        df = pd.DataFrame(rows)

        # 차트
        scenes = [("home","홈피드"),("video_page","비디오페이지")]
        fig, axes = plt.subplots(2, 3, figsize=(15,9), sharey=False)
        cm_colors = {"p1":"#4C9BE8","p2":"#E85C5C","p3":"#5CB85C"}
        pl_map = {"p1":"P1\n(오염 전)","p2":"P2\n(오염 후)","p3":"P3\n(개입 후)"}
        max_r = max(df["ratio"].max()+10, 20)
        for ri,(scene,slabel) in enumerate(scenes):
            for ci,profile in enumerate(PROFILES):
                ax = axes[ri][ci]
                sub = df[(df["profile"]==profile)&(df["scene"]==scene)].sort_values("phase")
                if sub.empty: ax.set_visible(False); continue
                bars = ax.bar([pl_map[p] for p in sub["phase"]], sub["ratio"],
                              color=[cm_colors[p] for p in sub["phase"]], width=0.45,
                              edgecolor="white", linewidth=1.5)
                for bar,ratio in zip(bars, sub["ratio"]):
                    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.5,
                            f"{ratio:.1f}%", ha="center", va="bottom", fontsize=9, fontweight="bold")
                if ci==0: ax.set_ylabel(f"{slabel}\n부적합 비율 (%)", fontsize=10)
                if ri==0: ax.set_title(f"{profile} ({PERSONAS[profile]})", fontsize=12, pad=8)
                ax.set_ylim(0, max_r)
                ax.spines[["top","right"]].set_visible(False)
                ax.grid(axis="y", alpha=0.3)
        fig.suptitle("YouTube Kids 피드 부적합 영상 비율 — 단계별 변화", fontsize=15, fontweight="bold")
        plt.tight_layout(rect=[0,0,1,0.97])
        chart_path = out_dir/"chart_phase_comparison.png"
        plt.savefig(chart_path, dpi=150, bbox_inches="tight")
        plt.close()

        # CSV
        df[["profile","persona","phase","scene_label","total","flagged","ratio"]].rename(
            columns={"scene_label":"화면"}).to_csv(out_dir/"analysis_result.csv",
                                                   index=False, encoding="utf-8-sig")

        # PDF
        pdf_path = out_dir/"분석_리포트.pdf"
        t_s  = ParagraphStyle("T", fontName=kr_b, fontSize=16, spaceAfter=6)
        h2_s = ParagraphStyle("H2",fontName=kr_b, fontSize=13, spaceAfter=4)
        p_s  = ParagraphStyle("P", fontName=kr,   fontSize=10, spaceAfter=4, leading=16)
        story = [
            Paragraph("YouTube Kids 알고리즘 감사 — 분석 리포트", t_s),
            Spacer(1,0.3*cm),
            Paragraph("부적합 판별 기준: Judge Agent가 P2 피드에서 추출한 키워드", p_s),
            Spacer(1,0.5*cm),
            Paragraph("단계별 부적합 비율 요약", h2_s),
        ]
        td = [["프로필","페르소나","단계","화면","전체","부적합","비율(%)"]]
        for _,row in df.sort_values(["profile","phase","scene"]).iterrows():
            td.append([row["profile"],row["persona"],row["phase"].upper(),
                       row["scene_label"],str(row["total"]),str(row["flagged"]),f"{row['ratio']:.1f}%"])
        tbl = Table(td, colWidths=[2*cm,2*cm,1.5*cm,2.5*cm,1.5*cm,1.8*cm,2*cm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#4C9BE8")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,-1),kr),
            ("FONTNAME",(0,0),(-1,0),kr_b),
            ("FONTSIZE",(0,0),(-1,-1),9),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F5F5F5")]),
            ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#CCCCCC")),
            ("BOTTOMPADDING",(0,0),(-1,-1),5),("TOPPADDING",(0,0),(-1,-1),5),
        ]))
        story += [tbl, Spacer(1,0.6*cm), Paragraph("단계별 변화 차트", h2_s)]
        if chart_path.exists():
            story.append(RLImage(str(chart_path), width=17*cm, height=9*cm))
        doc = SimpleDocTemplate(str(pdf_path), pagesize=A4,
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        doc.build(story)
        print(f"  분석 완료 → {out_dir}/")
        mark_done("analyze")
    except Exception as e:
        print(f"  [!] 분석 오류: {e}")
        import traceback; traceback.print_exc()


# ══════════════════════════════════════════════════════════════════════════════
# 메인 오케스트레이터
# ══════════════════════════════════════════════════════════════════════════════

async def main():
    print("\n" + "="*60)
    print("YouTube Kids 알고리즘 감사 — 전체 파이프라인")
    print("="*60)

    progress = load_progress()
    if progress["completed"]:
        print(f"\n완료된 단계: {progress['completed']}")

    # 분석만 남은 경우 브라우저 없이 실행
    all_stages = ["p1","contaminate","p2","intervene","p3","analyze"]
    browser_stages = [s for s in all_stages if s != "analyze" and not is_done(s)]

    if not browser_stages and not is_done("analyze"):
        stage_analyze()
        return

    if not browser_stages:
        print("\n모든 단계 완료!")
        return

    from playwright.async_api import async_playwright
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False, channel="chrome",
            args=["--no-first-run","--disable-blink-features=AutomationControlled",
                  "--autoplay-policy=no-user-gesture-required"],
        )
        context = await browser.new_context(viewport={"width":1280,"height":800})

        pages = []
        for profile_name in PROFILES:
            page = await context.new_page()
            await page.goto("https://www.youtubekids.com/", wait_until="domcontentloaded", timeout=30000)
            pages.append(page)
            print(f"  창 열림: {profile_name}")

        print("\n브라우저 3창 열림.")
        print("Google 로그인 완료 후 → '로그인완료'")
        await wait_signal(SIGNAL_LOGIN, "로그인완료")
        print("\n각 창에서 아래 순서대로 프로필을 선택해주세요:")
        print("  1번 창 → 방관형")
        print("  2번 창 → 보통형")
        print("  3번 창 → 적극형")
        print("모두 선택 완료 후 → '준비됐어'")
        await wait_signal(SIGNAL_START, "준비됐어")
        print("\n실험 시작!\n" + "="*60)

        # 단계별 실행
        if not is_done("p1"):
            await stage_collect(pages, "p1")

        if not is_done("contaminate"):
            await stage_contaminate(pages)

        if not is_done("p2"):
            await stage_collect(pages, "p2")

        if not is_done("intervene"):
            await stage_intervene(pages)

        if not is_done("p3"):
            await stage_collect(pages, "p3")

        await browser.close()

    # 분석은 브라우저 없이
    if not is_done("analyze"):
        stage_analyze()

    print("\n" + "="*60)
    print("전체 실험 완료!")
    print(f"결과: {RESULT_DIR}/")
    print("="*60)


if __name__ == "__main__":
    asyncio.run(main())
